"""
Google Sheets Template Generator für PISA-Befragungen

Erstellt Excel-Templates die in Google Sheets geöffnet werden können
mit automatischen Formeln zur Auswertung.
"""

from typing import List, Dict
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


def create_excel_template(
    scale_name: str,
    scale_title: str,
    items: List[Dict],
    pisa_average: float = 2.5
) -> BytesIO:
    """
    Erstellt Excel-Template mit automatischer Auswertung.

    Args:
        scale_name: Skalen-Code (z.B. "ANXMAT")
        scale_title: Deutscher Titel
        items: Liste von Items
        pisa_average: PISA Deutschland Durchschnitt für diese Skala

    Returns:
        BytesIO object mit Excel-Datei
    """

    wb = Workbook()

    # ===================================
    # BLATT 1: ROHDATEN
    # ===================================

    ws_data = wb.active
    ws_data.title = "Rohdaten"

    # Header styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column headers
    headers = ["Zeitstempel", "Schüler Name"]
    for idx, item in enumerate(items, 1):
        variable = item.get('variable_name', f'Item_{idx}')
        headers.append(f"{variable}")

    for col_idx, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Column widths
    ws_data.column_dimensions['A'].width = 20  # Zeitstempel
    ws_data.column_dimensions['B'].width = 25  # Name
    for col_idx in range(3, len(headers) + 1):
        ws_data.column_dimensions[get_column_letter(col_idx)].width = 15

    # Add example rows with formulas
    ws_data.cell(row=2, column=1, value="2024-01-15 10:30")
    ws_data.cell(row=2, column=2, value="Beispiel Schüler 1")
    for col_idx in range(3, len(headers) + 1):
        ws_data.cell(row=2, column=col_idx, value="")

    # Freeze panes
    ws_data.freeze_panes = "C2"

    # ===================================
    # BLATT 2: AUSWERTUNG
    # ===================================

    ws_eval = wb.create_sheet("Auswertung")

    # Title
    ws_eval.merge_cells('A1:F1')
    title_cell = ws_eval['A1']
    title_cell.value = f"📊 Auswertung: {scale_title}"
    title_cell.font = Font(size=16, bold=True, color="4472C4")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Info
    ws_eval['A3'] = "Skala:"
    ws_eval['B3'] = scale_name
    ws_eval['A4'] = "Anzahl Items:"
    ws_eval['B4'] = len(items)
    ws_eval['A5'] = "PISA DE Durchschnitt:"
    ws_eval['B5'] = pisa_average

    # Style info section
    for row in range(3, 6):
        ws_eval.cell(row=row, column=1).font = Font(bold=True)

    ws_eval['A7'] = ""  # Spacer

    # Headers für Auswertungstabelle
    eval_headers = ["Schüler", "Durchschnitt", "Vergleich zu PISA", "Status", "Risiko?"]
    for col_idx, header in enumerate(eval_headers, 1):
        cell = ws_eval.cell(row=8, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Column widths
    ws_eval.column_dimensions['A'].width = 25
    ws_eval.column_dimensions['B'].width = 15
    ws_eval.column_dimensions['C'].width = 18
    ws_eval.column_dimensions['D'].width = 15
    ws_eval.column_dimensions['E'].width = 12

    # Formula rows (will be populated when data is added)
    # Row 9 onwards: one row per student with formulas

    # Example formula row
    ws_eval['A9'] = '=Rohdaten!B2'  # Name from data sheet

    # Average formula (average of all item columns in Rohdaten)
    num_items = len(items)
    items_start_col = 3  # Column C in Rohdaten
    items_end_col = items_start_col + num_items - 1
    ws_eval['B9'] = f'=AVERAGE(Rohdaten!{get_column_letter(items_start_col)}2:{get_column_letter(items_end_col)}2)'

    # Comparison to PISA
    ws_eval['C9'] = f'=B9-$B$5'

    # Status with conditional color
    ws_eval['D9'] = '=IF(C9>0,"Über PISA","Unter PISA")'

    # Risiko flag
    ws_eval['E9'] = '=IF(B9<2.0,"⚠️ JA","")'

    # Add conditional formatting colors
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Number formats
    for row in range(9, 50):  # Prepare 50 rows
        ws_eval.cell(row=row, column=2).number_format = '0.00'
        ws_eval.cell(row=row, column=3).number_format = '+0.00;-0.00;0.00'

    # ===================================
    # BLATT 3: DASHBOARD
    # ===================================

    ws_dash = wb.create_sheet("Dashboard")

    # Title
    ws_dash.merge_cells('A1:F1')
    title_cell = ws_dash['A1']
    title_cell.value = f"📈 Dashboard: {scale_title}"
    title_cell.font = Font(size=16, bold=True, color="4472C4")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Metrics
    ws_dash['A3'] = "📊 Klassendurchschnitt:"
    ws_dash['B3'] = '=AVERAGE(Auswertung!B9:B50)'
    ws_dash['B3'].number_format = '0.00'
    ws_dash['B3'].font = Font(size=14, bold=True)

    ws_dash['A4'] = "🎯 PISA Deutschland:"
    ws_dash['B4'] = pisa_average
    ws_dash['B4'].number_format = '0.00'
    ws_dash['B4'].font = Font(size=14)

    ws_dash['A5'] = "📈 Differenz:"
    ws_dash['B5'] = '=B3-B4'
    ws_dash['B5'].number_format = '+0.00;-0.00;0.00'
    ws_dash['B5'].font = Font(size=14, bold=True)

    ws_dash['A7'] = "⚠️ Risikoschüler (Durchschnitt < 2.0):"
    ws_dash['B7'] = '=COUNTIF(Auswertung!B9:B50,"<2.0")'
    ws_dash['B7'].font = Font(size=14, bold=True, color="DC3545")

    # Styling
    for row in range(3, 8):
        ws_dash.cell(row=row, column=1).font = Font(bold=True)

    # Column widths
    ws_dash.column_dimensions['A'].width = 35
    ws_dash.column_dimensions['B'].width = 15

    # ===================================
    # BLATT 4: ANLEITUNG
    # ===================================

    ws_help = wb.create_sheet("Anleitung")

    instructions = [
        ("A1", "📋 ANLEITUNG - So verwendest du dieses Template", Font(size=14, bold=True, color="4472C4")),
        ("A3", "1️⃣ Google Sheets Setup", Font(bold=True, size=12)),
        ("A4", "   • Lade diese Datei in Google Drive hoch", None),
        ("A5", "   • Öffne mit Google Sheets", None),
        ("A6", "   • Kopiere die Web-App URL aus dem HTML-Formular", None),
        ("A8", "2️⃣ Datensammlung", Font(bold=True, size=12)),
        ("A9", "   • Teile den QR-Code oder Link mit deinen Schülern", None),
        ("A10", "   • Schüler füllen das Formular auf ihren Handys aus", None),
        ("A11", "   • Daten erscheinen automatisch im Tab 'Rohdaten'", None),
        ("A13", "3️⃣ Auswertung", Font(bold=True, size=12)),
        ("A14", "   • Tab 'Auswertung': Siehe individuelle Schüler-Ergebnisse", None),
        ("A15", "   • Tab 'Dashboard': Siehe Klassen-Überblick", None),
        ("A16", "   • Rote Markierungen: Schüler unter Durchschnitt", None),
        ("A18", "4️⃣ Interpretation", Font(bold=True, size=12)),
        ("A19", f"   • Werte > {pisa_average}: Überdurchschnittlich (im Vergleich zu PISA)", None),
        ("A20", f"   • Werte < {pisa_average}: Unterdurchschnittlich", None),
        ("A21", "   • Werte < 2.0: Risikogruppe (dringender Handlungsbedarf)", None),
        ("A23", "5️⃣ Handlungsempfehlungen", Font(bold=True, size=12)),
        ("A24", "   • Identifiziere Schüler mit niedrigen Werten", None),
        ("A25", "   • Entwickle gezielte Interventionen", None),
        ("A26", "   • Führe Follow-up Befragungen durch", None),
        ("A28", "💡 Tipp: Diese Skalen basieren auf PISA 2022 - wissenschaftlich validiert!", Font(italic=True)),
    ]

    for cell_ref, text, font_style in instructions:
        cell = ws_dash[cell_ref] if 'ws_dash' in locals() else ws_help[cell_ref]
        ws_help[cell_ref] = text
        if font_style:
            ws_help[cell_ref].font = font_style

    ws_help.column_dimensions['A'].width = 70

    # ===================================
    # SAVE TO BYTESIO
    # ===================================

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file


def create_google_apps_script_template(scale_name: str, items: List[Dict]) -> str:
    """
    Generiert Google Apps Script Code zum Einfügen in Google Sheets.

    Args:
        scale_name: Skalen-Code
        items: Liste von Items

    Returns:
        JavaScript code als String
    """

    item_columns = [item.get('variable_name', f'Item_{idx}') for idx, item in enumerate(items, 1)]

    script = f"""
// Google Apps Script für {scale_name}
// Dieses Script empfängt Daten vom HTML-Formular und speichert sie in Google Sheets

function doPost(e) {{
  try {{
    // Parse incoming data
    const data = JSON.parse(e.postData.contents);

    // Get the spreadsheet and sheet
    const ss = SpreadsheetApp.getActiveSpreadsheet();
    const sheet = ss.getSheetByName('Rohdaten');

    // Prepare row data
    const rowData = [
      data.timestamp || new Date().toISOString(),
      data.student_name || 'Unbekannt'
    ];

    // Add item responses in order
    const itemCols = {item_columns};
    itemCols.forEach(col => {{
      rowData.push(data[col] || '');
    }});

    // Append row
    sheet.appendRow(rowData);

    // Return success
    return ContentService.createTextOutput(JSON.stringify({{
      status: 'success',
      message: 'Daten gespeichert'
    }})).setMimeType(ContentService.MimeType.JSON);

  }} catch (error) {{
    return ContentService.createTextOutput(JSON.stringify({{
      status: 'error',
      message: error.toString()
    }})).setMimeType(ContentService.MimeType.JSON);
  }}
}}

// Test function
function testScript() {{
  const testData = {{
    postData: {{
      contents: JSON.stringify({{
        timestamp: new Date().toISOString(),
        student_name: 'Test Schüler',
        {', '.join([f'{col}: "3"' for col in item_columns[:3]])}
      }})
    }}
  }};

  const result = doPost(testData);
  Logger.log(result.getContent());
}}
"""

    return script
